"""
Seeds the India_GHG_Factors source into lci_processes, parsed directly from
GHG_Calculator_RECTIFIED_v6.xlsx (DB_Master + GHG_Master_Calculator sheets).

Deliberately does NOT hardcode label/unit text from memory -- it extracts
process_name and reference_unit directly from the GHG_Master_Calculator
sheet by finding the row whose VLOOKUP formula references each DB_Master
key. This was a real mistake caught during earlier work on this dataset
(a first pass used recalled labels and 17/60 were wrong) -- extracting
directly from the source avoids repeating it.
"""
import re

import openpyxl

from app.database import Base, SessionLocal, engine
from app.matching import EMBEDDING_MODEL, embed_text
from app.models import LciProcess
from app.units import normalize_unit

XLSX_PATH = "data/GHG_Calculator_RECTIFIED_v6.xlsx"


def classify_status(source_note: str) -> str:
    if "PLACEHOLDER" in source_note:
        return "placeholder"
    if "uplift" in source_note or "EPA GHG Equivalencies" in source_note:
        return "uplifted"
    if "Cross-validated" in source_note or "RECONCILED" in source_note:
        return "proxy"
    return "clean"


def extract_labels_and_units(wb) -> dict[str, tuple[str, str]]:
    """For every DB_Master key, find its real label + unit from the actual
    calculator sheet rows that use it, rather than guessing."""
    ws = wb["GHG_Master_Calculator"]
    extracted: dict[str, tuple[str, str]] = {}
    for row in ws.iter_rows():
        for cell in row:
            if isinstance(cell.value, str) and cell.value.startswith("=VLOOKUP"):
                match = re.search(r'"([^"]+)"', cell.value)
                if not match:
                    continue
                key = match.group(1)
                label = ws.cell(row=cell.row, column=1).value
                unit = ws.cell(row=cell.row, column=3).value
                if key not in extracted:  # first usage wins; keys can repeat across categories
                    extracted[key] = (label, unit)
    return extracted


def load_factors(path: str):
    # data_only=False is required here -- with data_only=True, formula cells
    # return their cached computed value instead of the formula text, so the
    # "=VLOOKUP(...)" string match in extract_labels_and_units would never
    # fire. DB_Master's own Key/Factor/Source columns are plain literals
    # (not formulas), so this doesn't affect reading them correctly.
    wb = openpyxl.load_workbook(path, data_only=False)
    labels = extract_labels_and_units(wb)
    ws = wb["DB_Master"]
    factors = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        key, factor, note = row[0], row[1], row[2]
        if not key:
            continue
        label, unit = labels.get(key, (key, "unit"))
        factors.append({
            "key": key,
            "factor": factor,
            "source": note or "",
            "status": classify_status(note or ""),
            "label": label,
            "unit": unit,
        })
    return factors


def seed():
    Base.metadata.create_all(bind=engine)
    factors = load_factors(XLSX_PATH)

    db = SessionLocal()
    try:
        existing = db.query(LciProcess).filter(
            LciProcess.database_source == "India_GHG_Factors"
        ).count()
        if existing > 0:
            print(f"India_GHG_Factors already seeded ({existing} rows); skipping.")
            return

        missing_label = [f["key"] for f in factors if not f["label"] or f["label"] == f["key"]]
        missing_unit = [f["key"] for f in factors if not f["unit"] or f["unit"] == "unit" and f["key"] != "Laptop_Unit"]
        if missing_label:
            print(f"WARNING: {len(missing_label)} keys had no extracted label, fell back to key name: {missing_label}")

        for f in factors:
            text = f"{f['label']} {f['source']}"
            db.add(LciProcess(
                process_uuid=f["key"],
                database_source="India_GHG_Factors",
                database_version="v6",
                process_name=f["label"],
                reference_product=f["label"],
                reference_unit=normalize_unit(str(f["unit"])),
                geography="IN",
                system_model="Direct Factor",
                description=f["source"],
                emission_factor=float(f["factor"]),
                emission_factor_source=f["source"],
                data_quality_status=f["status"],
                embedding=embed_text(text),
                embedding_model=EMBEDDING_MODEL,
            ))
        db.commit()
        print(f"Seeded {len(factors)} India_GHG_Factors rows.")
        status_counts = {}
        for f in factors:
            status_counts[f["status"]] = status_counts.get(f["status"], 0) + 1
        print(f"Provenance breakdown: {status_counts}")
    finally:
        db.close()


if __name__ == "__main__":
    seed()
